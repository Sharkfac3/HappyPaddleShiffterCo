"""
recalc.py — Verifies BOM.xlsx has zero formula errors.
Scans every cell in every sheet for Excel error values.
"""

import sys
from openpyxl import load_workbook

ERROR_VALUES = {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}

path = r"C:\Users\sharkfac3\Workspace\coding\HappyPaddleShiffterCo\BOM.xlsx"
wb = load_workbook(path, data_only=True)

errors_found = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().upper() in ERROR_VALUES:
                errors_found.append(
                    f"  Sheet '{sheet_name}' cell {cell.coordinate}: {val!r}"
                )

if errors_found:
    print(f"FAIL — {len(errors_found)} formula error(s) found:")
    for e in errors_found:
        print(e)
    sys.exit(1)
else:
    sheets = wb.sheetnames
    total_cells = sum(
        ws.max_row * ws.max_column
        for ws in (wb[s] for s in sheets)
    )
    print(f"PASS — Zero formula errors in {len(sheets)} sheets ({total_cells} cells checked).")
    print(f"Sheets: {', '.join(sheets)}")
    sys.exit(0)
