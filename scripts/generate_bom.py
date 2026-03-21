"""
generate_bom.py — Creates BOM.xlsx for HappyPaddleShifterCo
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Helpers ────────────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, italic=False, size=11, color="000000"):
    return Font(name="Arial", bold=bold, italic=italic, size=size, color=color)

def apply_header_row(ws, row, cols, text, font, fill, merge=True, alignment=None):
    """Write a merged header row."""
    first = get_column_letter(1) + str(row)
    last  = get_column_letter(cols) + str(row)
    if merge:
        ws.merge_cells(f"{first}:{last}")
    cell = ws[first]
    cell.value = text
    cell.font  = font
    cell.fill  = fill
    cell.alignment = alignment or Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color

def freeze_rows(ws, row=4):
    ws.freeze_panes = f"A{row}"

def write_row(ws, row_num, values, font=None, fill=None, wrap=True):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def write_cell(ws, row, col, value, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:  c.font  = font
    if fill:  c.fill  = fill
    c.alignment = align or Alignment(horizontal="left", vertical="top", wrap_text=True)
    return c

def alt_fill(row_num, even_fill, odd_fill):
    # row_num is 1-based; data starts at offset
    return even_fill if row_num % 2 == 0 else odd_fill

WHITE_FILL  = hex_fill("FFFFFF")
LGREY_FILL  = hex_fill("F2F2F2")
LRED_FILL   = hex_fill("FFE0E0")
WARN_RED    = hex_fill("FFC7CE")
LYELLOW     = hex_fill("FFFF00")
LGREEN      = hex_fill("E2EFDA")
LAVENDER    = hex_fill("F3EEFF")
LORANGE     = hex_fill("FCE4D6")

# ============================================================
# SHEET 1 — Electronics
# ============================================================
ws1 = wb.active
ws1.title = "Electronics"
set_tab_color(ws1, "4472C4")

# Row heights
ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 14

DARK_BLUE1 = hex_fill("1F3864")
HDR_BLUE   = hex_fill("2E4057")
WHITE_TXT  = "FFFFFF"

# Row 1
apply_header_row(ws1, 1, 7,
    "HappyPaddleShifterCo — Bill of Materials",
    make_font(bold=True, size=14, color=WHITE_TXT),
    DARK_BLUE1)

# Row 2
apply_header_row(ws1, 2, 7,
    "Jeep XJ Cherokee AW4 Paddle Shifter Controller",
    make_font(italic=True, size=11, color=WHITE_TXT),
    DARK_BLUE1)

# Row 3
apply_header_row(ws1, 3, 7,
    "Last Verified: 2026-03-20  |  Source: ArduinoCode/SYSTEM.md and .agents/skills/hardware-bom/SKILL.md",
    make_font(size=9, color=WHITE_TXT),
    DARK_BLUE1)

# Row 4 — column headers
headers = ["Qty", "Description", "Specification", "Part Number",
           "Unit Cost (USD)", "Notes / Warnings", "Status"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_BLUE
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data rows 5-9
rows_data = [
    (5, [1, "Arduino Mega 2560",
         "ATmega2560 MCU, 54 digital I/O, 16 analog, 5V logic, 256KB flash",
         "—", "—",
         "Original or compatible clone. Do NOT use Uno — insufficient I/O pins and flash memory.",
         "Confirmed"]),
    (6, [1, "OLED Display",
         "WaveShare 1.5\" RGB OLED, SSD1351 driver, 128×128 px, 3.3V VCC",
         "—", "—",
         "3.3V VCC ONLY. Applying 5V to VCC permanently destroys this display.",
         "Confirmed — WARNING: See note"]),
    (7, [1, "Solenoid Driver Board",
         "Relay or high-side driver, minimum 3 channels, 12V coil-compatible",
         "—", "—",
         "NEVER connect solenoids directly to Arduino pins. Channels needed: S1, S2, SLU.",
         "Confirmed"]),
    (8, [3, "Flyback Diode",
         "1N4007, 1A, 1000V PIV",
         "1N4007", "—",
         "One across each solenoid coil (S1, S2, SLU). Prevents back-EMF damage. MANDATORY.",
         "Confirmed"]),
    (9, [2, "Paddle Switch",
         "Omron D2JW-01K11, SPDT N/O momentary, straight lever, IP67, 30VDC/100mA, chassis mount",
         "Omron D2JW-01K11", "—",
         "IP67 waterproof. 100,000 electrical cycle life. NOT interchangeable with D2F-5L (pocket depth differs by 1.6 mm).",
         "Confirmed"]),
]

for row_num, vals in rows_data:
    fill = WHITE_FILL if (row_num % 2 == 1) else LGREY_FILL
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=row_num, column=col, value=val)
        c.font = make_font()
        # notes column (col 6) gets red fill for rows 6,7,8
        if col == 6 and row_num in (6, 7, 8):
            c.fill = LRED_FILL
        else:
            c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Row heights for data rows
for r in range(5, 10):
    ws1.row_dimensions[r].height = 60

# Row 10: blank separator
ws1.row_dimensions[10].height = 8

# Row 11: safety warning label
c11 = ws1.cell(row=11, column=1, value="Safety Warnings — Mandatory")
c11.font = make_font(bold=True, size=11)
ws1.merge_cells("A11:G11")

warn_rows = {
    12: "DISPLAY: 3.3V VCC ONLY. Connecting 5V permanently destroys the SSD1351 OLED.",
    13: "SOLENOIDS: Run on 12V, draw up to 2A. Use relay/driver board only. Install 1N4007 flyback diode across each coil.",
    14: "NSS: Continuity switch only — NOT powered. Do not apply 12V to NSS pins. Use INPUT_PULLUP; wire common leg to GND.",
}
for r, txt in warn_rows.items():
    ws1.merge_cells(f"A{r}:G{r}")
    c = ws1.cell(row=r, column=1, value=txt)
    c.font  = make_font(bold=False, size=10)
    c.fill  = WARN_RED
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws1.row_dimensions[r].height = 30

set_col_widths(ws1, [8, 28, 40, 22, 18, 55, 14])
freeze_rows(ws1, 4)

# ============================================================
# SHEET 2 — Vehicle Connector
# ============================================================
ws2 = wb.create_sheet("Vehicle Connector")
set_tab_color(ws2, "70AD47")

DK_GREEN  = hex_fill("1E4620")
HDR_GREEN = hex_fill("375623")

apply_header_row(ws2, 1, 5,
    "AW4 Neutral Safety Switch (NSS) Connector",
    make_font(bold=True, size=14, color=WHITE_TXT),
    DK_GREEN)

apply_header_row(ws2, 2, 5,
    "Choose connector by vehicle model year — Jeep XJ Cherokee 1987–2001",
    make_font(size=10, color=WHITE_TXT),
    DK_GREEN)

# Row 3: warning
ws2.merge_cells("A3:E3")
c3 = ws2["A3"]
c3.value = "WARNING: The NSS is a continuity switch ONLY — not powered. Never apply 12V to NSS pins."
c3.font  = make_font(bold=True, color="C00000")
c3.fill  = LYELLOW
c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[3].height = 20

# Row 4: headers
hdr4 = ["Year Range", "Connector Description", "Omix-ADA Part No.", "OEM Part No.", "Notes / Status"]
for col, h in enumerate(hdr4, 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_GREEN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row 5
r5_vals = ["1987–1996",
           "8-pin Deutsch connector, pins labelled A–H",
           "17216.03", "83503712",
           "Pinout confirmed from 1993 FSM. Pins: A=common/GND, B-C=Park/Neutral, E=Reverse, G=3rd hold, H=1-2 hold."]
for col, val in enumerate(r5_vals, 1):
    c = ws2.cell(row=5, column=col, value=val)
    c.font  = make_font()
    c.fill  = LGREEN
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws2.row_dimensions[5].height = 50

# Row 6
r6_vals = ["1997–2001",
           "Different physical connector housing",
           "4882173", "4882173",
           "Continuity map identical to 1987-1996. Physical pin label-to-wire mapping UNVERIFIED — confirm with multimeter before wiring."]
for col, val in enumerate(r6_vals, 1):
    c = ws2.cell(row=6, column=col, value=val)
    c.font  = make_font()
    c.fill  = LYELLOW
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws2.row_dimensions[6].height = 50

# Row 7: blank
ws2.row_dimensions[7].height = 8

# Row 8: NSS label
ws2.merge_cells("A8:E8")
c8 = ws2["A8"]
c8.value = "NSS Continuity Map (consistent across all XJ years)"
c8.font  = make_font(bold=True, size=11)
c8.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[8].height = 18

# Row 9: sub-headers
hdr9 = ["Selector Position", "NSS Pins Closed", "Arduino Pin", "Signal", "Notes"]
for col, h in enumerate(hdr9, 1):
    c = ws2.cell(row=9, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_GREEN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[9].height = 18

# Rows 10-15
nss_rows = [
    (10, ["Park",      "B ↔ C",       "D4",  "Park/Neutral",  "Electrically indistinguishable from Neutral"]),
    (11, ["Neutral",   "B ↔ C",       "D4",  "Park/Neutral",  "Electrically indistinguishable from Park"]),
    (12, ["Reverse",   "A ↔ E",       "D5",  "Reverse",       "Pin A is shared common — wire A to GND"]),
    (13, ["Drive",     "(none close)", "—",   "Drive",         "Detected by ALL four monitored pins reading HIGH simultaneously"]),
    (14, ["3rd Hold",  "A ↔ G",       "D6",  "3rd Hold",      "Pin A is shared common — wire A to GND"]),
    (15, ["1-2 Hold",  "A ↔ H",       "D7",  "1-2 Hold",      "Pin A is shared common — wire A to GND"]),
]
for row_num, vals in nss_rows:
    fill = WHITE_FILL if (row_num % 2 == 0) else LGREY_FILL
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row_num, column=col, value=val)
        c.font  = make_font()
        c.fill  = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws2.row_dimensions[row_num].height = 30

set_col_widths(ws2, [14, 38, 22, 16, 60])
freeze_rows(ws2, 4)

# ============================================================
# SHEET 3 — Paddle Switch Detail
# ============================================================
ws3 = wb.create_sheet("Paddle Switch Detail")
set_tab_color(ws3, "7030A0")

DK_PURPLE  = hex_fill("3E1154")
HDR_PURPLE = hex_fill("4B2067")
SEC_PURPLE = hex_fill("7030A0")

apply_header_row(ws3, 1, 3,
    "Paddle Switch — Omron D2JW-01K11",
    make_font(bold=True, size=14, color=WHITE_TXT),
    DK_PURPLE)

apply_header_row(ws3, 2, 3,
    "Selected switch for HappyPaddleShifterCo paddle hardware",
    make_font(italic=True, color=WHITE_TXT),
    DK_PURPLE)

apply_header_row(ws3, 3, 3,
    "Source: Omron D2JW-01K11 datasheet, verified 2026-03-20",
    make_font(size=9, color=WHITE_TXT),
    DK_PURPLE)

# Row 4 headers
hdr3 = ["Section", "Parameter", "Value"]
for col, h in enumerate(hdr3, 1):
    c = ws3.cell(row=4, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_PURPLE
    c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[4].height = 18

# Row 5: section label
ws3.merge_cells("A5:C5")
c5 = ws3["A5"]
c5.value = "ELECTRICAL SPECIFICATIONS"
c5.font  = make_font(bold=True, color=WHITE_TXT)
c5.fill  = SEC_PURPLE
c5.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[5].height = 18

elec_rows = [
    (6,  ["Electrical", "Contact Configuration", "SPDT (normally-open + normally-closed contacts)"]),
    (7,  ["Electrical", "Voltage Rating",         "30 VDC"]),
    (8,  ["Electrical", "Current Rating",          "100 mA DC"]),
    (9,  ["Electrical", "Operating Force",         "82 gf"]),
    (10, ["Electrical", "Release Force",           "16 gf"]),
    (11, ["Electrical", "Electrical Life",         "100,000 cycles (~13.7 years at 20 shifts/day)"]),
    (12, ["Electrical", "Mechanical Life",         "1,000,000 cycles"]),
    (13, ["Electrical", "Operating Temperature",   "−40°C to +85°C"]),
    (14, ["Electrical", "IP Rating",               "IP67 — dust-tight and fully waterproof"]),
]
for row_num, vals in elec_rows:
    fill = WHITE_FILL if (row_num % 2 == 0) else LAVENDER
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row_num, column=col, value=val)
        c.font  = make_font()
        c.fill  = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3.row_dimensions[row_num].height = 18

# Row 15: section label
ws3.merge_cells("A15:C15")
c15 = ws3["A15"]
c15.value = "MECHANICAL / MOUNTING"
c15.font  = make_font(bold=True, color=WHITE_TXT)
c15.fill  = SEC_PURPLE
c15.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[15].height = 18

mech_rows = [
    (16, ["Mechanical", "Actuator Type",                    "Lever, Straight"]),
    (17, ["Mechanical", "Body Width",                       "12.7 mm"]),
    (18, ["Mechanical", "Body Height",                      "12.3 mm"]),
    (19, ["Mechanical", "Body Depth",                       "5.3 mm ± 0.1 mm"]),
    (20, ["Mechanical", "Mounting Hole Diameter",           "2.35 mm"]),
    (21, ["Mechanical", "Mounting Hole Spacing",            "3.95 × 3.95 mm"]),
    (22, ["Mechanical", "Lever Arc Radius",                 "R16.5 mm (stainless steel t0.3 mm)"]),
    (23, ["Mechanical", "Operating Position (lever tip travel)", "8.4 mm ± 0.8 mm"]),
    (24, ["Mechanical", "Pretravel",                        "6.4 mm max"]),
    (25, ["Mechanical", "Overtravel",                       "1.4 mm min"]),
    (26, ["Mechanical", "Movement Differential",            "0.7 mm max"]),
    (27, ["Mechanical", "Termination",                      "Solder lug — short wire leads to PCB/Arduino"]),
    (28, ["Mechanical", "Mount Type",                       "Chassis mount — screws directly to 3D-printed paddle body"]),
]
for row_num, vals in mech_rows:
    fill = WHITE_FILL if (row_num % 2 == 0) else LAVENDER
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row_num, column=col, value=val)
        c.font  = make_font()
        c.fill  = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3.row_dimensions[row_num].height = 18

# Row 29: blank
ws3.row_dimensions[29].height = 8

# Row 30: 3D print note
ws3.merge_cells("A30:C30")
c30 = ws3["A30"]
c30.value = ("3D PRINT NOTE: Switch pocket must place lever tip 8.4 mm from paddle contact surface. "
             "This is 1.6 mm DEEPER than the former D2F-5L pocket (6.8 mm). "
             "The D2JW-01K11 and D2F-5L are NOT dimensionally interchangeable.")
c30.font  = make_font(bold=True, size=10)
c30.fill  = LYELLOW
c30.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws3.row_dimensions[30].height = 50

set_col_widths(ws3, [18, 38, 45])
freeze_rows(ws3, 4)

# ============================================================
# SHEET 4 — Arduino Libraries
# ============================================================
ws4 = wb.create_sheet("Arduino Libraries")
set_tab_color(ws4, "A5A5A5")

DK_GREY  = hex_fill("404040")
HDR_GREY = hex_fill("595959")

apply_header_row(ws4, 1, 4,
    "Required Arduino Libraries",
    make_font(bold=True, size=14, color=WHITE_TXT),
    DK_GREY)

apply_header_row(ws4, 2, 4,
    "No purchase required — install via Arduino IDE Library Manager",
    make_font(italic=True, color=WHITE_TXT),
    DK_GREY)

# Row 3 headers
hdr4_cols = ["Library", "Source", "How to Install", "Notes"]
for col, h in enumerate(hdr4_cols, 1):
    c = ws4.cell(row=3, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_GREY
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws4.row_dimensions[3].height = 18

lib_rows = [
    (4, ["Adafruit_SSD1351",
         "Arduino Library Manager",
         'Sketch → Include Library → Manage Libraries → search "SSD1351" → Install',
         "OLED display driver for SSD1351 chip"]),
    (5, ["Adafruit_GFX",
         "Arduino Library Manager",
         'Sketch → Include Library → Manage Libraries → search "Adafruit GFX" → Install',
         "Graphics primitives library — required dependency of Adafruit_SSD1351"]),
    (6, ["SPI",
         "Arduino IDE built-in",
         "No install required — included with Arduino IDE",
         "Hardware SPI. On Mega 2560: pin 51 = MOSI (DIN), pin 52 = SCK (CLK). These pins are FIXED and cannot be reassigned."]),
]
for row_num, vals in lib_rows:
    fill = WHITE_FILL if (row_num % 2 == 0) else LGREY_FILL
    for col, val in enumerate(vals, 1):
        c = ws4.cell(row=row_num, column=col, value=val)
        c.font  = make_font()
        c.fill  = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.row_dimensions[row_num].height = 50

set_col_widths(ws4, [22, 28, 60, 55])
freeze_rows(ws4, 3)

# NOTE: freeze_rows(ws4, 3) freezes A3, meaning rows 1 & 2 are frozen.
# The spec says freeze first 3 rows → freeze_panes = A4
ws4.freeze_panes = "A4"

# ============================================================
# SHEET 5 — Pin Reference
# ============================================================
ws5 = wb.create_sheet("Pin Reference")
set_tab_color(ws5, "ED7D31")

DK_ORANGE  = hex_fill("833C00")
HDR_ORANGE = hex_fill("A9441D")

apply_header_row(ws5, 1, 6,
    "Arduino Mega 2560 — Pin Assignments",
    make_font(bold=True, size=14, color=WHITE_TXT),
    DK_ORANGE)

apply_header_row(ws5, 2, 6,
    "Source of truth: ArduinoCode/SYSTEM.md — do not modify this table independently of the firmware spec",
    make_font(italic=True, color=WHITE_TXT),
    DK_ORANGE)

# Row 3: warning
ws5.merge_cells("A3:F3")
c3b = ws5["A3"]
c3b.value = "WARNING: Solenoid pins (11, 12, 13) must connect via relay/driver board ONLY — never directly."
c3b.font  = make_font(bold=True, color="C00000")
c3b.fill  = LYELLOW
c3b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws5.row_dimensions[3].height = 20

# Row 4: headers
hdr5 = ["Arduino Pin", "Direction", "Mode", "Connected To", "Signal Group", "Notes"]
for col, h in enumerate(hdr5, 1):
    c = ws5.cell(row=4, column=col, value=h)
    c.font  = make_font(bold=True, color=WHITE_TXT)
    c.fill  = HDR_ORANGE
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws5.row_dimensions[4].height = 18

pin_rows = [
    (5,  ["2",  "INPUT",  "INPUT_PULLUP", "Shift Up paddle → GND",           "Paddle Input",    "Active LOW. Triggers on falling edge. 50ms debounce."]),
    (6,  ["3",  "INPUT",  "INPUT_PULLUP", "Shift Down paddle → GND",         "Paddle Input",    "Active LOW. Triggers on falling edge. 50ms debounce."]),
    (7,  ["4",  "INPUT",  "INPUT_PULLUP", "NSS Pin B (Park/Neutral) → GND",  "NSS Input",       "Active LOW. Park and Neutral are electrically indistinguishable."]),
    (8,  ["5",  "INPUT",  "INPUT_PULLUP", "NSS Pin E (Reverse) → GND",       "NSS Input",       "Active LOW. Pin A is common — wire NSS Pin A to GND."]),
    (9,  ["6",  "INPUT",  "INPUT_PULLUP", "NSS Pin G (3rd hold) → GND",      "NSS Input",       "Active LOW. Pin A is common — wire NSS Pin A to GND."]),
    (10, ["7",  "INPUT",  "INPUT_PULLUP", "NSS Pin H (1-2 hold) → GND",      "NSS Input",       "Active LOW. Pin A is common — wire NSS Pin A to GND."]),
    (11, ["11", "OUTPUT", "—",            "S1 Solenoid → relay/driver board CH1",  "Solenoid Output", "WARNING: Via driver board ONLY. Never connect solenoid directly to Arduino pin."]),
    (12, ["12", "OUTPUT", "—",            "S2 Solenoid → relay/driver board CH2",  "Solenoid Output", "WARNING: Via driver board ONLY. Never connect solenoid directly to Arduino pin."]),
    (13, ["13", "OUTPUT", "—",            "SLU Solenoid → relay/driver board CH3", "Solenoid Output", "WARNING: Via driver board ONLY. Never connect solenoid directly to Arduino pin."]),
    (14, ["51", "OUTPUT", "Hardware MOSI (SPI)", "Display DIN",               "Display (SPI)",   "Fixed hardware SPI MOSI on Mega 2560. Cannot be reassigned to another pin."]),
    (15, ["52", "OUTPUT", "Hardware SCK (SPI)",  "Display CLK",               "Display (SPI)",   "Fixed hardware SPI SCK on Mega 2560. Cannot be reassigned to another pin."]),
    (16, ["A3", "OUTPUT", "—",            "Display RES",                      "Display Control", "—"]),
    (17, ["A4", "OUTPUT", "—",            "Display DC",                       "Display Control", "—"]),
    (18, ["A5", "OUTPUT", "—",            "Display CS",                       "Display Control", "—"]),
]

for row_num, vals in pin_rows:
    # solenoid rows get red fill
    if row_num in (11, 12, 13):
        row_fill = LRED_FILL
    else:
        row_fill = WHITE_FILL if (row_num % 2 == 1) else LORANGE

    for col, val in enumerate(vals, 1):
        c = ws5.cell(row=row_num, column=col, value=val)
        c.font  = make_font()
        c.fill  = row_fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws5.row_dimensions[row_num].height = 40

set_col_widths(ws5, [14, 12, 22, 42, 20, 55])
freeze_rows(ws5, 4)

# ============================================================
# Save
# ============================================================
out_path = r"C:\Users\sharkfac3\Workspace\coding\HappyPaddleShiffterCo\BOM.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
